"""Excel (XLSX) 文档解析器 — 输出 Markdown 表格格式"""

import io
from typing import Any, Dict, List

from openpyxl import load_workbook

from app.services.parser.base import BaseParser

# 最大读取行数，防止超大 Excel OOM
MAX_ROWS = 100_000


class ExcelParser(BaseParser):
    """使用 openpyxl 提取 Excel 内容，输出 Markdown 表格格式"""

    def parse(self, data: bytes, filename: str) -> Dict[str, Any]:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts: List[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            row_count = 0

            for row in ws.iter_rows(values_only=True):
                if row_count >= MAX_ROWS:
                    break
                cells = [str(c) if c is not None else "" for c in row]
                # 跳过全空行
                if not any(c.strip() for c in cells):
                    continue
                rows.append(cells)
                row_count += 1

            if not rows:
                continue

            total_rows += len(rows)

            # 输出为 Markdown 表格
            sheet_md = f"## {sheet_name}\n\n"
            sheet_md += _rows_to_markdown_table(rows)

            if row_count >= MAX_ROWS:
                sheet_md += f"\n\n> ⚠️ 该 sheet 超过 {MAX_ROWS} 行，仅显示前 {MAX_ROWS} 行。"

            parts.append(sheet_md)

        wb.close()

        metadata: Dict[str, Any] = {
            "parse_method": "openpyxl",
            "output_format": "markdown",
            "sheets": wb.sheetnames,
            "total_rows": total_rows,
        }

        return {
            "text": "\n\n".join(parts),
            "pages": len(wb.sheetnames),
            "metadata": metadata,
        }


def _rows_to_markdown_table(rows: List[List[str]]) -> str:
    """将行数据转为 Markdown 表格"""
    if not rows:
        return ""

    # 统一列数（取最大列数）
    max_cols = max(len(r) for r in rows)
    for r in rows:
        while len(r) < max_cols:
            r.append("")

    # 第一行作为表头
    header = rows[0]
    separator = ["---"] * max_cols
    lines = [
        "| " + " | ".join(h.replace("|", "\\|") for h in header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in rows[1:]:
        lines.append("| " + " | ".join(c.replace("|", "\\|") for c in row) + " |")

    return "\n".join(lines)
